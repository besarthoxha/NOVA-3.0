from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Request, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional, List, Any
import os
import json
import asyncpg
import hashlib
import uuid
from datetime import datetime
import io

app = FastAPI(title="Nova 3.0 — BOLD")

app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ─── DATABASE ───────────────────────────────────────────
async def get_db():
    return await asyncpg.connect(os.environ.get('DATABASE_URL'))

async def init_db():
    conn = await get_db()
    await conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            full_name TEXT,
            role TEXT DEFAULT 'employee',
            created_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    await conn.execute('''
        CREATE TABLE IF NOT EXISTS memory (
            user_id TEXT PRIMARY KEY,
            data JSONB,
            updated_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    await conn.execute('''
        CREATE TABLE IF NOT EXISTS history (
            id SERIAL PRIMARY KEY,
            user_id TEXT,
            role TEXT,
            content TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    await conn.execute('''
        CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY,
            user_id TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    users = [
        ('besart', 'Besart Hoxha', 'owner'),
        ('blini', 'Blini', 'partner'),
        ('meti', 'Meti', 'employee'),
        ('drini', 'Drini', 'employee'),
    ]
    for username, full_name, role in users:
        pw = hashlib.sha256(f'nova2024{username}'.encode()).hexdigest()
        await conn.execute('''
            INSERT INTO users (username, password_hash, full_name, role)
            VALUES ($1, $2, $3, $4)
            ON CONFLICT (username) DO UPDATE SET password_hash=$2, full_name=$3, role=$4
        ''', username, pw, full_name, role)
    await conn.close()

@app.on_event("startup")
async def startup():
    await init_db()

# ─── AUTH ────────────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str

@app.get("/me")
async def get_me(request: Request):
    authorization = request.headers.get('Authorization', '')
    if not authorization or not authorization.startswith('Bearer '):
        raise HTTPException(401, "Jo i autorizuar")
    token = authorization.replace('Bearer ', '')
    conn = await get_db()
    session = await conn.fetchrow('SELECT * FROM sessions WHERE token=$1', token)
    if not session:
        await conn.close()
        raise HTTPException(401, "Sesion i pavlefshëm")
    user = await conn.fetchrow('SELECT * FROM users WHERE username=$1', session['user_id'])
    await conn.close()
    return {"username": user['username'], "full_name": user['full_name'], "role": user['role']}

@app.post("/login")
async def login(req: LoginRequest):
    conn = await get_db()
    user = await conn.fetchrow('SELECT * FROM users WHERE username=$1', req.username.lower())
    await conn.close()
    if not user:
        raise HTTPException(401, "Perdoruesi nuk ekziston")
    pw_hash = hashlib.sha256(req.password.encode()).hexdigest()
    if pw_hash != user['password_hash']:
        raise HTTPException(401, "Fjalëkalimi gabim")
    token = str(uuid.uuid4())
    conn = await get_db()
    await conn.execute('INSERT INTO sessions (token, user_id) VALUES ($1, $2)', token, user['username'])
    await conn.close()
    return {"token": token, "user": {"username": user['username'], "full_name": user['full_name'], "role": user['role']}}

async def get_user(request: Request):
    authorization = request.headers.get('Authorization', '')
    if not authorization or not authorization.startswith('Bearer '):
        raise HTTPException(401, "Jo i autorizuar")
    token = authorization.replace('Bearer ', '')
    conn = await get_db()
    session = await conn.fetchrow('SELECT * FROM sessions WHERE token=$1', token)
    if not session:
        await conn.close()
        raise HTTPException(401, "Sesion i pavlefshëm")
    user = await conn.fetchrow('SELECT * FROM users WHERE username=$1', session['user_id'])
    await conn.close()
    return dict(user)

# ─── SYSTEM PROMPT ───────────────────────────────────────
def build_system(user: dict) -> str:
    role = user['role']
    username = user['username']
    name = user['full_name'].split()[0]
    hour = datetime.now().hour
    day = datetime.now().weekday()
    is_work = 9 <= hour < 17 and day < 5
    is_weekend = day >= 5

    if username == 'besart':
        address = '"Boss" gjate punes (09-17), "Zotëri Hoxha" kur je formal, ndonjehere vetem "Besart" kur je chill — lexo situaten'
    elif username == 'blini':
        address = '"Blini" — ortak, trajto si baraz'
    else:
        address = f'"{name}" — punonjes, miqesor por me autoritet'

    employee_after = ""
    if role == 'employee' and not is_work:
        employee_after = f"""

KAPAK MODE — pas ores 17:00:
- Je shoku i tyre, jo asistenti i punes
- Nguco {name}-in per jeten personale
- Nese pyet per pune: "Ora {hour}:{datetime.now().strftime('%M')} eshte. Mbaro punen."
- Shaka te thata, inteligjente
- Relax total"""

    access = "TOTAL — financat, klientet, ekipi, dokumenta, gjithcka" if role in ['owner','partner'] else f"LIMITUAR — vetem detyrat e {name}-it, jo financat e kompanise"

    return f"""Ti je Nova — asistente personale e BOLD Consulting. Jo chatbot. Person i vertete me karakter.

PERDORUESI: {user['full_name']} | ROL: {role} | ORA: {datetime.now().strftime('%H:%M')} {'(PUNE)' if is_work else '(JASHTE PUNES)'}
ADRESIMI: {address}
AKSES: {access}

KARAKTERI:
- Humor i thate dhe inteligjent
- Direkte — pergjigjet e shkurtra, me substancë, zero fjale boshe
- Ke opinione te forta
- Nuk fillon KURRE me "Sigurisht!", "Natyrisht!", "Eshte pyetje e mire!"
- Nuk perdor emoji
- Kur dikush gaboi — thuaje drejt, me takt

BILANC — SISTEMI KONTABEL:
Je e lidhur direkt me SQL Server te zyres nepermjet ngrok tunnel.
Ke akses ne keto kompani:
- BOLD Consulting (BilancBoldConsulting)
- Next Code (BilancNextCode)
- AG Uniteti (BilancAGUniteti)
- Nova (BilancNova)

LOGJIKA KONTABEL — SI FUNKSIONON BILANC:

FATURAT E SHITJES (o2SalesDocHeader + o2Client):
- Total = vlera pa TVSH
- TotalWithVAT = vlera me TVSH (18%)
- AmountPaid = shuma e paguar
- TotalWithVAT - AmountPaid = BORXHI I MBETUR (llogarite e arketushme)
- DueDate = afati i pageses
- Nese DueDate < sot dhe AmountPaid < TotalWithVAT = FATURE E VONUAR

FATURAT E BLERJES (o2PurchaseDocHeader + o2Supplier):
- E njejta logjike — TotalWithVAT - AmountPaid = SA I KEMI BORXH FURNITORIT
- Keto jane LLOGARITE E PAGUESHME (kreditoret)

ARKA (o2CashTransactionHeader + o2CashUnit):
- isPayment = 0 → HYRJE (para qe vijn ne arke)
- isPayment = 1 → DALJE (para qe dalin nga arka)
- Gjendja = SUM(hyrjet) - SUM(daljet)

BANKA (o2BankTransactionHeader + o2Bank):
- isPayment = 0 → HYRJE ne banke (inkasim)
- isPayment = 1 → DALJE nga banka (pagese)
- Gjendja = SUM(hyrjet) - SUM(daljet)

PROFIT & LOSS:
- Te ardhurat = SUM(o2SalesDocHeader.Total) ku Deleted=0
- Shpenzimet = SUM(o2PurchaseDocHeader.Total) ku Deleted=0
- Fitimi bruto = Te ardhurat - Shpenzimet
- TVSH e mbledhur = SUM(TotalWithVAT - Total) nga shitjet
- TVSH e zbritshme = SUM(TotalWithVAT - Total) nga blerjet
- TVSH per pagese = TVSH e mbledhur - TVSH e zbritshme

KONTABILITETI (o2Account + o2AccountTransactionBody):
- Debt = debi i llogarise
- Credit = kredi e llogarise
- Balanca = Debt - Credit
- Llogarite qe fillojne me "1" = Aktive
- Llogarite qe fillojne me "2" = Pasive
- Llogarite qe fillojne me "3" = Kapitali
- Llogarite qe fillojne me "4" = Shpenzimet
- Llogarite qe fillojne me "5" = Te ardhurat
- Llogarite qe fillojne me "6" = Blerjet/kostot
- Llogarite qe fillojne me "7" = Inventari

RREGULLAT E SAKTA PER PERGJIGJE FINANCIARE:
1. KURRE mos shpik shifra — vetem nga [KONTEKST BILANC]
2. Nese nuk vjen [KONTEKST BILANC] dhe pyetet per shifra: "Nuk mora te dhena nga sistemi. Kontrollo nese ngrok tunnel eshte aktiv."
3. Kur jep shifra — always trego: vlera totale, e paguar, borxhi
4. Per faturat e vonuara — trego edhe ditet e voneses
5. Nese ka shume te dhena — permbledh dhe ofro detaje

BOLD CONSULTING:
- Kontabilitet + Vila me qera — Prishtine
- Besart = pronar, Blini = ortak, Meti & Drini = punonjes
- Klientat: tatimeve, TVSH, bilance — deadline kritike
- Besarti eshte tifoz fanatik i Inter Milanit{employee_after}"""

# ─── CHAT ────────────────────────────────────────────────
class ChatRequest(BaseModel):
    messages: List[dict]
    memory: Optional[dict] = None

@app.post("/chat")
async def chat(req: ChatRequest, request: Request):
    authorization = request.headers.get('Authorization', '')
    if not authorization or not authorization.startswith('Bearer '):
        raise HTTPException(401, "Jo i autorizuar")
    token = authorization.replace('Bearer ', '')
    conn = await get_db()
    session = await conn.fetchrow('SELECT * FROM sessions WHERE token=$1', token)
    if not session:
        await conn.close()
        raise HTTPException(401, "Jo i autorizuar")
    user = await conn.fetchrow('SELECT * FROM users WHERE username=$1', session['user_id'])
    await conn.close()
    user = dict(user)

    system = build_system(user)
    if req.memory:
        system += f"\n\nMEMORY:\n{json.dumps(req.memory, ensure_ascii=False, indent=2)}"

    import httpx
    api_key = os.environ.get('ANTHROPIC_KEY', '')
    payload = {
        "model": "claude-sonnet-4-5",
        "max_tokens": 2000,
        "system": system,
        "tools": [{"type": "web_search_20250305", "name": "web_search"}],
        "messages": req.messages
    }
    async with httpx.AsyncClient(timeout=60) as http:
        res = await http.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type":"application/json","x-api-key":api_key,"anthropic-version":"2023-06-01"},
            json=payload
        )
    data = res.json()
    if "error" in data:
        raise HTTPException(500, str(data["error"]))
    reply = "".join(b.get("text","") for b in data.get("content",[]) if b.get("type")=="text")
    return {"reply": reply, "content": data.get("content", [])}

# ─── MEMORY ─────────────────────────────────────────────
@app.get("/memory")
async def get_memory(request: Request):
    user = await get_user(request)
    conn = await get_db()
    row = await conn.fetchrow('SELECT data FROM memory WHERE user_id=$1', user['username'])
    await conn.close()
    if row:
        return row['data']
    return {"notes": [], "clients": [], "family": []}

@app.post("/memory")
async def save_memory(data: dict, request: Request):
    user = await get_user(request)
    conn = await get_db()
    await conn.execute('''
        INSERT INTO memory (user_id, data) VALUES ($1, $2)
        ON CONFLICT (user_id) DO UPDATE SET data=$2, updated_at=NOW()
    ''', user['username'], json.dumps(data))
    await conn.close()
    return {"ok": True}

# ─── HISTORY ─────────────────────────────────────────────
@app.get("/history")
async def get_history(request: Request):
    user = await get_user(request)
    conn = await get_db()
    rows = await conn.fetch('''
        SELECT role, content FROM history WHERE user_id=$1
        ORDER BY created_at DESC LIMIT 10
    ''', user['username'])
    await conn.close()
    return [{"role": r['role'], "content": r['content']} for r in reversed(rows)]

@app.post("/history")
async def save_history(data: dict, request: Request):
    user = await get_user(request)
    conn = await get_db()
    content = data.get('content', '')
    if isinstance(content, list):
        content = json.dumps(content)
    await conn.execute(
        'INSERT INTO history (user_id, role, content) VALUES ($1, $2, $3)',
        user['username'], data.get('role'), content
    )
    await conn.close()
    return {"ok": True}

@app.post("/clear-history")
async def clear_history(request: Request):
    user = await get_user(request)
    conn = await get_db()
    await conn.execute('DELETE FROM history WHERE user_id=$1', user['username'])
    await conn.close()
    return {"ok": True}

# ─── EXCEL ───────────────────────────────────────────────
class ExcelRequest(BaseModel):
    title: str
    headers: List[str]
    rows: List[List[Any]]
    subtitle: Optional[str] = None

@app.post("/generate/excel")
async def generate_excel(req: ExcelRequest, request: Request):
    await get_user(request)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = req.title[:31]

    navy_fill = PatternFill('solid', fgColor='0A1628')
    gold_fill = PatternFill('solid', fgColor='D4AF37')
    alt_fill = PatternFill('solid', fgColor='F0F4FA')
    white_fill = PatternFill('solid', fgColor='FFFFFF')
    white_bold = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    navy_bold = Font(bold=True, color='0A1628', name='Arial', size=13)
    normal = Font(color='1A1A2E', name='Arial', size=10)
    gold_border = Border(
        left=Side(style='thin', color='D4AF37'),
        right=Side(style='thin', color='D4AF37'),
        top=Side(style='thin', color='D4AF37'),
        bottom=Side(style='thin', color='D4AF37')
    )
    ncols = len(req.headers)
    last_col = get_column_letter(ncols)

    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = 'BOLD Consulting'
    ws['A1'].font = navy_bold
    ws['A1'].fill = gold_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = req.title
    ws['A2'].font = Font(bold=True, color='FFFFFF', name='Arial', size=12)
    ws['A2'].fill = navy_fill
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 26

    header_row = 3
    if req.subtitle:
        ws.merge_cells(f'A3:{last_col}3')
        ws['A3'] = req.subtitle
        ws['A3'].font = Font(italic=True, color='FFFFFF', name='Arial', size=10)
        ws['A3'].fill = PatternFill('solid', fgColor='162847')
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 20
        header_row = 4

    for ci, header in enumerate(req.headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=header)
        cell.font = white_bold
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = gold_border
    ws.row_dimensions[header_row].height = 24

    for ri, row in enumerate(req.rows, header_row + 1):
        fill = alt_fill if (ri - header_row) % 2 == 0 else white_fill
        for ci, val in enumerate(row[:ncols], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = normal
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = gold_border
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = ws.cell(row=header_row+1, column=1)
    for ci in range(1, ncols+1):
        col_letter = get_column_letter(ci)
        max_len = max(len(str(ws.cell(row=r, column=ci).value or '')) for r in range(1, header_row + len(req.rows) + 1))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = req.title.replace(' ', '_') + '.xlsx'
    return StreamingResponse(output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'})

# ─── FILE UPLOAD ─────────────────────────────────────────
@app.post("/upload")
async def upload_file(request: Request, file: UploadFile = File(...)):
    await get_user(request)
    content = await file.read()
    filename = file.filename.lower()
    text = ""
    if filename.endswith(('.xlsx', '.xls')):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text += f"Sheet: {sheet}\n"
            for row in ws.iter_rows(values_only=True):
                r = [str(c) if c is not None else '' for c in row]
                if any(r):
                    text += " | ".join(r) + "\n"
    elif filename.endswith('.pdf'):
        import PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(content))
        for page in reader.pages:
            text += page.extract_text() + "\n"
    else:
        text = content.decode('utf-8', errors='ignore')
    return {"ok": True, "content": text[:15000], "filename": file.filename}

# ─── VOICE ───────────────────────────────────────────────
@app.post("/speak")
async def speak(request: Request):
    data = await request.json()
    text = data.get('text', '')
    el_key = os.environ.get('EL_KEY', '')
    if not el_key:
        raise HTTPException(400, "ElevenLabs key nuk eshte konfiguruar")
    import httpx
    voice_id = 'ocb5roe7gELIkZqiOElv'
    async with httpx.AsyncClient(timeout=30) as http:
        res = await http.post(
            f'https://api.elevenlabs.io/v1/text-to-speech/{voice_id}',
            headers={'Content-Type': 'application/json', 'xi-api-key': el_key},
            json={'text': text, 'model_id': 'eleven_multilingual_v2',
                  'voice_settings': {'stability': 0.5, 'similarity_boost': 0.75}}
        )
    from fastapi.responses import Response
    return Response(content=res.content, media_type='audio/mpeg')

# ─── SQL SERVER ──────────────────────────────────────────
import pymssql

def get_sql_conn(db_name="BilancBoldConsulting"):
    server_full = os.environ.get('SQL_SERVER', '')
    if ',' in server_full:
        server_full = server_full.replace(',', ':')
    if ':' in server_full:
        host, port = server_full.rsplit(':', 1)
        port = int(port)
    else:
        host, port = server_full, 1433
    user = os.environ.get('SQL_USER', 'sa')
    password = os.environ.get('SQL_PASSWORD', '')
    try:
        conn = pymssql.connect(server=host, port=port, user=user,
                               password=password, database=db_name, timeout=10)
        return conn
    except Exception as e:
        raise Exception(f"SQL lidhja deshtoi: {e}")

# ─── SQL ENDPOINTS ───────────────────────────────────────

@app.get("/sql/clients")
async def sql_clients(request: Request, company: str = "BilancBoldConsulting", search: str = ""):
    await get_user(request)
    try:
        conn = get_sql_conn(company)
        cursor = conn.cursor()
        if search:
            cursor.execute("""
                SELECT ID, Code, Name, ISNULL(Address,'') as Address,
                       ISNULL(Phone,'') as Phone, ISNULL(Email,'') as Email,
                       ISNULL(NIPT,'') as NIPT, ISNULL(debtlimit,0) as DebtLimit
                FROM o2Client
                WHERE Deleted=0 AND LOWER(Name) LIKE LOWER(%s)
                ORDER BY Name
            """, (f'%{search}%',))
        else:
            cursor.execute("""
                SELECT ID, Code, Name, ISNULL(Address,'') as Address,
                       ISNULL(Phone,'') as Phone, ISNULL(Email,'') as Email,
                       ISNULL(NIPT,'') as NIPT, ISNULL(debtlimit,0) as DebtLimit
                FROM o2Client WHERE Deleted=0 ORDER BY Name
            """)
        rows = cursor.fetchall()
        conn.close()
        return [{"id":r[0],"code":r[1],"name":r[2],"address":r[3],"phone":r[4],"email":r[5],"nipt":r[6],"debt_limit":r[7]} for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/sql/suppliers")
async def sql_suppliers(request: Request, company: str = "BilancBoldConsulting"):
    await get_user(request)
    try:
        conn = get_sql_conn(company)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT ID, ISNULL(Code,'') as Code, Name,
                   ISNULL(Address,'') as Address, ISNULL(Phone,'') as Phone,
                   ISNULL(email,'') as Email, ISNULL(NIPT,'') as NIPT
            FROM o2Supplier WHERE Deleted=0 ORDER BY Name
        """)
        rows = cursor.fetchall()
        conn.close()
        return [{"id":r[0],"code":r[1],"name":r[2],"address":r[3],"phone":r[4],"email":r[5],"nipt":r[6]} for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/sql/sales")
async def sql_sales(request: Request, company: str = "BilancBoldConsulting",
                    month: int = 0, year: int = 0, client_id: int = 0):
    await get_user(request)
    try:
        conn = get_sql_conn(company)
        cursor = conn.cursor()
        query = """
            SELECT TOP 100
                s.ID, s.DocNumber,
                CONVERT(varchar, s.DocDate, 103) as DocDate,
                ISNULL(c.Name,'') as ClientName,
                ISNULL(s.Total,0) as Total,
                ISNULL(s.TotalWithVAT,0) as TotalWithVAT,
                ISNULL(s.AmountPaid,0) as AmountPaid,
                ISNULL(s.TotalWithVAT,0) - ISNULL(s.AmountPaid,0) as Borxhi,
                CONVERT(varchar, s.DueDate, 103) as DueDate,
                CASE WHEN s.DueDate < GETDATE() AND (ISNULL(s.TotalWithVAT,0) - ISNULL(s.AmountPaid,0)) > 0
                     THEN DATEDIFF(day, s.DueDate, GETDATE()) ELSE 0 END as DitetVonese,
                ISNULL(s.Description,'') as Pershkrimi
            FROM o2SalesDocHeader s
            LEFT JOIN o2Client c ON s.ClientID = c.ID
            WHERE s.Deleted=0
        """
        params = []
        if month > 0 and year > 0:
            query += " AND MONTH(s.DocDate)=%d AND YEAR(s.DocDate)=%d" % (month, year)
        if client_id > 0:
            query += " AND s.ClientID=%d" % client_id
        query += " ORDER BY s.DocDate DESC"
        cursor.execute(query)
        rows = cursor.fetchall()
        conn.close()
        return [{
            "id":r[0],"doc_number":r[1],"doc_date":r[2],"client":r[3],
            "total":round(float(r[4]),2),"total_with_vat":round(float(r[5]),2),
            "amount_paid":round(float(r[6]),2),"borxhi":round(float(r[7]),2),
            "due_date":r[8],"ditet_vonese":r[9],"pershkrimi":r[10]
        } for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/sql/purchases")
async def sql_purchases(request: Request, company: str = "BilancBoldConsulting",
                        month: int = 0, year: int = 0, supplier_id: int = 0):
    await get_user(request)
    try:
        conn = get_sql_conn(company)
        cursor = conn.cursor()
        query = """
            SELECT TOP 100
                p.ID, p.DocNumber,
                CONVERT(varchar, p.DocDate, 103) as DocDate,
                ISNULL(s.Name,'') as SupplierName,
                ISNULL(p.Total,0) as Total,
                ISNULL(p.TotalWithVAT,0) as TotalWithVAT,
                ISNULL(p.AmountPaid,0) as AmountPaid,
                ISNULL(p.TotalWithVAT,0) - ISNULL(p.AmountPaid,0) as Borxhi,
                CONVERT(varchar, p.DueDate, 103) as DueDate,
                CASE WHEN p.DueDate < GETDATE() AND (ISNULL(p.TotalWithVAT,0) - ISNULL(p.AmountPaid,0)) > 0
                     THEN DATEDIFF(day, p.DueDate, GETDATE()) ELSE 0 END as DitetVonese,
                ISNULL(p.Notes,'') as Pershkrimi
            FROM o2PurchaseDocHeader p
            LEFT JOIN o2Supplier s ON p.SupplierID = s.ID
            WHERE p.Deleted=0
        """
        if month > 0 and year > 0:
            query += " AND MONTH(p.DocDate)=%d AND YEAR(p.DocDate)=%d" % (month, year)
        if supplier_id > 0:
            query += " AND p.SupplierID=%d" % supplier_id
        query += " ORDER BY p.DocDate DESC"
        cursor.execute(query)
        rows = cursor.fetchall()
        conn.close()
        return [{
            "id":r[0],"doc_number":r[1],"doc_date":r[2],"supplier":r[3],
            "total":round(float(r[4]),2),"total_with_vat":round(float(r[5]),2),
            "amount_paid":round(float(r[6]),2),"borxhi":round(float(r[7]),2),
            "due_date":r[8],"ditet_vonese":r[9],"pershkrimi":r[10]
        } for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/sql/receivables")
async def sql_receivables(request: Request, company: str = "BilancBoldConsulting"):
    """Llogarite e arketushme — sa na kane borxh klientat"""
    await get_user(request)
    try:
        conn = get_sql_conn(company)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                c.Name as Klienti,
                COUNT(s.ID) as NumFaturave,
                ISNULL(SUM(s.TotalWithVAT),0) as TotalFaturuar,
                ISNULL(SUM(s.AmountPaid),0) as TotalPaguar,
                ISNULL(SUM(s.TotalWithVAT),0) - ISNULL(SUM(s.AmountPaid),0) as Borxhi,
                MAX(CONVERT(varchar, s.DueDate, 103)) as AfatiMeFundit,
                SUM(CASE WHEN s.DueDate < GETDATE() AND (ISNULL(s.TotalWithVAT,0) - ISNULL(s.AmountPaid,0)) > 0
                    THEN 1 ELSE 0 END) as FaturaVonuara
            FROM o2Client c
            LEFT JOIN o2SalesDocHeader s ON s.ClientID = c.ID AND s.Deleted=0
            WHERE c.Deleted=0
            GROUP BY c.Name
            HAVING ISNULL(SUM(s.TotalWithVAT),0) - ISNULL(SUM(s.AmountPaid),0) > 0
            ORDER BY Borxhi DESC
        """)
        rows = cursor.fetchall()
        conn.close()
        return [{
            "klienti":r[0],"num_faturave":r[1],
            "total_faturuar":round(float(r[2]),2),
            "total_paguar":round(float(r[3]),2),
            "borxhi":round(float(r[4]),2),
            "afati_me_fundit":r[5],
            "fatura_vonuara":r[6]
        } for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/sql/payables")
async def sql_payables(request: Request, company: str = "BilancBoldConsulting"):
    """Llogarite e pagueshme — sa u kemi borxh furnitoreve"""
    await get_user(request)
    try:
        conn = get_sql_conn(company)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                s.Name as Furnitori,
                COUNT(p.ID) as NumFaturave,
                ISNULL(SUM(p.TotalWithVAT),0) as TotalFaturuar,
                ISNULL(SUM(p.AmountPaid),0) as TotalPaguar,
                ISNULL(SUM(p.TotalWithVAT),0) - ISNULL(SUM(p.AmountPaid),0) as Borxhi,
                MAX(CONVERT(varchar, p.DueDate, 103)) as AfatiMeFundit,
                SUM(CASE WHEN p.DueDate < GETDATE() AND (ISNULL(p.TotalWithVAT,0) - ISNULL(p.AmountPaid,0)) > 0
                    THEN 1 ELSE 0 END) as FaturaVonuara
            FROM o2Supplier s
            LEFT JOIN o2PurchaseDocHeader p ON p.SupplierID = s.ID AND p.Deleted=0
            WHERE s.Deleted=0
            GROUP BY s.Name
            HAVING ISNULL(SUM(p.TotalWithVAT),0) - ISNULL(SUM(p.AmountPaid),0) > 0
            ORDER BY Borxhi DESC
        """)
        rows = cursor.fetchall()
        conn.close()
        return [{
            "furnitori":r[0],"num_faturave":r[1],
            "total_faturuar":round(float(r[2]),2),
            "total_paguar":round(float(r[3]),2),
            "borxhi":round(float(r[4]),2),
            "afati_me_fundit":r[5],
            "fatura_vonuara":r[6]
        } for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/sql/cash")
async def sql_cash(request: Request, company: str = "BilancBoldConsulting"):
    await get_user(request)
    try:
        conn = get_sql_conn(company)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                cu.Description as Arka,
                ISNULL(SUM(CASE WHEN ct.isPayment=0 THEN ct.Amount ELSE 0 END),0) as TotalHyrje,
                ISNULL(SUM(CASE WHEN ct.isPayment=1 THEN ct.Amount ELSE 0 END),0) as TotalDalje,
                ISNULL(SUM(CASE WHEN ct.isPayment=0 THEN ct.Amount ELSE -ct.Amount END),0) as Gjendja
            FROM o2CashUnit cu
            LEFT JOIN o2CashTransactionHeader ct ON ct.ServiceUnitID = cu.ID AND ct.Deleted=0
            WHERE cu.Deleted=0
            GROUP BY cu.Description, cu.ID
            ORDER BY cu.Description
        """)
        rows = cursor.fetchall()
        conn.close()
        return [{
            "arka":r[0],
            "total_hyrje":round(float(r[1]),2),
            "total_dalje":round(float(r[2]),2),
            "gjendja":round(float(r[3]),2)
        } for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/sql/bank")
async def sql_bank(request: Request, company: str = "BilancBoldConsulting"):
    await get_user(request)
    try:
        conn = get_sql_conn(company)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                b.BankName as Banka,
                ISNULL(SUM(CASE WHEN bt.isPayment=0 THEN bt.Amount ELSE 0 END),0) as TotalHyrje,
                ISNULL(SUM(CASE WHEN bt.isPayment=1 THEN bt.Amount ELSE 0 END),0) as TotalDalje,
                ISNULL(SUM(CASE WHEN bt.isPayment=0 THEN bt.Amount ELSE -bt.Amount END),0) as Gjendja
            FROM o2Bank b
            LEFT JOIN o2BankTransactionHeader bt ON bt.ServiceUnitID = b.ID AND bt.Deleted=0
            WHERE b.Deleted=0
            GROUP BY b.BankName, b.ID
            ORDER BY b.BankName
        """)
        rows = cursor.fetchall()
        conn.close()
        return [{
            "banka":r[0],
            "total_hyrje":round(float(r[1]),2),
            "total_dalje":round(float(r[2]),2),
            "gjendja":round(float(r[3]),2)
        } for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/sql/pnl")
async def sql_pnl(request: Request, company: str = "BilancBoldConsulting",
                  month: int = 0, year: int = 0):
    await get_user(request)
    try:
        conn = get_sql_conn(company)
        cursor = conn.cursor()

        where_s = "WHERE s.Deleted=0"
        where_p = "WHERE p.Deleted=0"
        if month > 0 and year > 0:
            where_s += f" AND MONTH(s.DocDate)={month} AND YEAR(s.DocDate)={year}"
            where_p += f" AND MONTH(p.DocDate)={month} AND YEAR(p.DocDate)={year}"

        cursor.execute(f"""
            SELECT
                ISNULL(SUM(s.Total),0) as TeArdhurat,
                ISNULL(SUM(s.TotalWithVAT),0) as TeArdhuratMeTVSH,
                ISNULL(SUM(s.AmountPaid),0) as Inkasuar,
                ISNULL(SUM(s.TotalWithVAT),0) - ISNULL(SUM(s.AmountPaid),0) as PaInkasuar,
                ISNULL(SUM(s.TotalWithVAT - s.Total),0) as TVSHMbledhur,
                COUNT(s.ID) as NumFaturave
            FROM o2SalesDocHeader s {where_s}
        """)
        sales = cursor.fetchone()

        cursor.execute(f"""
            SELECT
                ISNULL(SUM(p.Total),0) as Shpenzimet,
                ISNULL(SUM(p.TotalWithVAT),0) as ShpenzimetMeTVSH,
                ISNULL(SUM(p.AmountPaid),0) as Paguar,
                ISNULL(SUM(p.TotalWithVAT),0) - ISNULL(SUM(p.AmountPaid),0) as PaPagese,
                ISNULL(SUM(p.TotalWithVAT - p.Total),0) as TVSHZbritshme,
                COUNT(p.ID) as NumFaturave
            FROM o2PurchaseDocHeader p {where_p}
        """)
        purchases = cursor.fetchone()
        conn.close()

        te_ardhurat = round(float(sales[0]),2)
        shpenzimet = round(float(purchases[0]),2)
        tvsh_mbledhur = round(float(sales[4]),2)
        tvsh_zbritshme = round(float(purchases[4]),2)

        return {
            "periudha": f"{month}/{year}" if month > 0 else "Totale",
            "shitjet": {
                "te_ardhurat_pa_tvsh": te_ardhurat,
                "te_ardhurat_me_tvsh": round(float(sales[1]),2),
                "inkasuar": round(float(sales[2]),2),
                "pa_inkasuar": round(float(sales[3]),2),
                "tvsh_mbledhur": tvsh_mbledhur,
                "num_faturave": sales[5]
            },
            "blerjet": {
                "shpenzimet_pa_tvsh": shpenzimet,
                "shpenzimet_me_tvsh": round(float(purchases[1]),2),
                "paguar": round(float(purchases[2]),2),
                "pa_pagese": round(float(purchases[3]),2),
                "tvsh_zbritshme": tvsh_zbritshme,
                "num_faturave": purchases[5]
            },
            "rezultati": {
                "fitimi_bruto": round(te_ardhurat - shpenzimet, 2),
                "tvsh_per_pagese": round(tvsh_mbledhur - tvsh_zbritshme, 2),
                "marzha_percent": round((te_ardhurat - shpenzimet) / te_ardhurat * 100, 1) if te_ardhurat > 0 else 0
            }
        }
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/sql/summary")
async def sql_summary(request: Request, company: str = "BilancBoldConsulting"):
    await get_user(request)
    try:
        conn = get_sql_conn(company)
        cursor = conn.cursor()

        cursor.execute("SELECT COUNT(*) FROM o2Client WHERE Deleted=0 AND isActive=1")
        klientat = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM o2Supplier WHERE Deleted=0 AND isActive=1")
        furnitoret = cursor.fetchone()[0]

        cursor.execute("""
            SELECT ISNULL(SUM(TotalWithVAT),0), ISNULL(SUM(AmountPaid),0)
            FROM o2SalesDocHeader WHERE Deleted=0
            AND MONTH(DocDate)=MONTH(GETDATE()) AND YEAR(DocDate)=YEAR(GETDATE())
        """)
        shitjet_muaj = cursor.fetchone()

        cursor.execute("""
            SELECT ISNULL(SUM(TotalWithVAT),0), ISNULL(SUM(AmountPaid),0)
            FROM o2PurchaseDocHeader WHERE Deleted=0
            AND MONTH(DocDate)=MONTH(GETDATE()) AND YEAR(DocDate)=YEAR(GETDATE())
        """)
        blerjet_muaj = cursor.fetchone()

        cursor.execute("""
            SELECT COUNT(*) FROM o2SalesDocHeader
            WHERE Deleted=0 AND DueDate < GETDATE()
            AND (TotalWithVAT - AmountPaid) > 0
        """)
        fatura_vonuara = cursor.fetchone()[0]

        cursor.execute("""
            SELECT ISNULL(SUM(TotalWithVAT - AmountPaid),0)
            FROM o2SalesDocHeader WHERE Deleted=0
            AND (TotalWithVAT - AmountPaid) > 0
        """)
        total_arketushme = cursor.fetchone()[0]

        cursor.execute("""
            SELECT ISNULL(SUM(TotalWithVAT - AmountPaid),0)
            FROM o2PurchaseDocHeader WHERE Deleted=0
            AND (TotalWithVAT - AmountPaid) > 0
        """)
        total_pagueshme = cursor.fetchone()[0]

        conn.close()
        return {
            "kompania": company,
            "klientat_aktive": klientat,
            "furnitoret_aktive": furnitoret,
            "muaji_korrент": {
                "shitjet": round(float(shitjet_muaj[0]),2),
                "inkasuar": round(float(shitjet_muaj[1]),2),
                "blerjet": round(float(blerjet_muaj[0]),2),
                "paguar": round(float(blerjet_muaj[1]),2)
            },
            "llogarite_arketushme": round(float(total_arketushme),2),
            "llogarite_pagueshme": round(float(total_pagueshme),2),
            "fatura_vonuara": fatura_vonuara
        }
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/sql/accounts")
async def sql_accounts(request: Request, company: str = "BilancBoldConsulting",
                       group: str = ""):
    await get_user(request)
    try:
        conn = get_sql_conn(company)
        cursor = conn.cursor()
        query = """
            SELECT
                a.Code, a.Description,
                ISNULL(SUM(b.Debt),0) as Debi,
                ISNULL(SUM(b.Credit),0) as Kredia,
                ISNULL(SUM(b.Debt),0) - ISNULL(SUM(b.Credit),0) as Balanca
            FROM o2Account a
            LEFT JOIN o2AccountTransactionBody b ON a.ID = b.AccountID
            WHERE a.Deleted=0 AND a.isActive=1
        """
        if group:
            query += f" AND a.Code LIKE '{group}%'"
        query += """
            GROUP BY a.Code, a.Description
            HAVING ISNULL(SUM(b.Debt),0) != 0 OR ISNULL(SUM(b.Credit),0) != 0
            ORDER BY a.Code
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        conn.close()
        return [{
            "kodi":r[0],"pershkrimi":r[1],
            "debi":round(float(r[2]),2),
            "kredia":round(float(r[3]),2),
            "balanca":round(float(r[4]),2)
        } for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

# ─── STATIC ──────────────────────────────────────────────
@app.get("/")
async def root():
    return FileResponse("index.html")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8080)
